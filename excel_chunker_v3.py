"""
Excel preprocessing & chunking pipeline for RAG (v3 — region-based).

Handles real-world messy Excel layouts:
- Multiple tables per sheet, each with own headers
- Free-text bullet-point paragraphs (merged full-width rows)
- Footnotes BETWEEN tables (not just at the bottom)
- Bare footnote markers: *, **, ***, not just (1), (**)
- Vertical merged cells (subtopic labels → forward-fill)
- Horizontal merged cells (section headers → hierarchy)
- Subtitle/label rows preceding tables ("Oevzub Rıüowyğly;")

Architecture: Parse sheet into typed REGIONS first, then chunk each region
appropriately. This avoids the v2 problem of assuming one table per sheet.

Dependencies: openpyxl
"""

import re
from dataclasses import dataclass, field
from enum import Enum
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

class RegionType(Enum):
    TITLE = "title"           # Wide merged section header
    TEXT = "text"             # Full-width merged paragraph (bullet points etc.)
    TABLE = "table"           # Tabular data with headers
    LABEL = "label"           # Subtitle/label row before a table
    FOOTNOTE = "footnote"     # Footnote definition row
    EMPTY = "empty"


@dataclass
class MergeInfo:
    value: str | None
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    @property
    def col_span(self) -> int:
        return self.max_col - self.min_col + 1

    @property
    def row_span(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def is_horizontal_header(self) -> bool:
        return self.row_span == 1 and self.col_span > 1

    @property
    def is_vertical_label(self) -> bool:
        return self.col_span == 1 and self.row_span > 1


@dataclass
class Region:
    rtype: RegionType
    start_row: int
    end_row: int
    text: str = ""
    rows_data: list = field(default_factory=list)   # for TABLE regions
    col_headers: dict = field(default_factory=dict)  # for TABLE regions
    header_row: int | None = None


@dataclass
class Chunk:
    text: str
    metadata: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Footnote patterns — now includes bare *, **, *** without parentheses
# ---------------------------------------------------------------------------

# Matches footnote markers IN data cells (inline references)
INLINE_MARKER_RE = re.compile(
    r'(\(\d+\))'           # (1), (12)
    r'|(\(\*+\))'          # (*), (**)
    r'|(\([!]+\))'         # (!), (!!)
    r'|(\[\d+\])'          # [1], [12]
    r'|(?<!\w)(\*{1,3})(?!\w|\*)'  # bare *, **, *** (not inside words)
    r'|([¹²³⁴⁵⁶⁷⁸⁹⁰†‡§])'
)

# Matches the START of a footnote definition row
# e.g.: "*Explanation text" or "(1) Explanation" or "** Some note"
FOOTNOTE_DEF_RE = re.compile(
    r'^\s*'
    r'(?:'
    r'\(\d+\)'              # (1), (12)
    r'|\(\*+\)'             # (*), (**)
    r'|\([!]+\)'            # (!), (!!)
    r'|\[\d+\]'             # [1], [12]
    r'|\*{1,3}(?!\*)'       # *, **, ***
    r'|[¹²³⁴⁵⁶⁷⁸⁹⁰†‡§]'
    r')'
    r'\s*\S'                # followed by actual text
)

# Extracts marker + text from a footnote definition
FOOTNOTE_EXTRACT_RE = re.compile(
    r'^\s*(\(\d+\)|\(\*+\)|\([!]+\)|\[\d+\]|\*{1,3}(?!\*)|[¹²³⁴⁵⁶⁷⁸⁹⁰†‡§])\s*(.+)',
    re.DOTALL,
)

FOOTNOTE_SECTION_KEYWORDS = re.compile(
    r'^\s*(not(?:e|es)|source|sources|reference|references|legend|footnot|açıklama|kaynak)',
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Step 1: Catalog & unmerge
# ---------------------------------------------------------------------------

def catalog_merges(ws) -> list[MergeInfo]:
    catalog = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        catalog.append(MergeInfo(
            value=ws.cell(row=min_row, column=min_col).value,
            min_row=min_row, max_row=max_row,
            min_col=min_col, max_col=max_col,
        ))
    return catalog


def unmerge_and_fill(ws, catalog: list[MergeInfo]) -> None:
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(mr))
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(mr))
        for row in ws.iter_rows(min_col=min_col, min_row=min_row,
                                max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = value


# ---------------------------------------------------------------------------
# Step 2: Classify each row
# ---------------------------------------------------------------------------

def _row_values(ws, row_idx: int) -> list:
    return [ws.cell(row=row_idx, column=c).value
            for c in range(ws.min_column, ws.max_column + 1)]


def _non_none_count(values: list) -> int:
    return sum(1 for v in values if v is not None)


def _get_row_text(ws, row_idx: int) -> str:
    """Combine all cell values in a row into one string, deduplicating unmerged repeats."""
    vals = [ws.cell(row=row_idx, column=c).value
            for c in range(ws.min_column, ws.max_column + 1)]
    # After unmerge, a merged cell's value is duplicated across all columns.
    # Deduplicate consecutive identical values.
    seen_prev = None
    deduped = []
    for v in vals:
        if v is not None and v != seen_prev:
            deduped.append(str(v))
        seen_prev = v
    return " ".join(deduped).strip()


def _is_full_width_merge(row_idx: int, catalog: list[MergeInfo], data_width: int,
                         threshold: float = 0.6) -> bool:
    """Check if row has a merge spanning most of the sheet width."""
    for m in catalog:
        if m.min_row <= row_idx <= m.max_row and m.col_span >= data_width * threshold:
            return True
    return False


def _is_footnote_row(text: str) -> bool:
    """Check if a row is a footnote definition."""
    if not text:
        return False
    return bool(FOOTNOTE_DEF_RE.match(text))


def _is_label_row(text: str) -> bool:
    """
    A label row is a short text (subtitle) that precedes a table.
    Usually ends with ; or : and is not a bullet point.
    """
    if not text or len(text) > 120:
        return False
    if text.startswith("●") or text.startswith("•") or text.startswith("-"):
        return False
    if text.rstrip().endswith(";") or text.rstrip().endswith(":"):
        return True
    return False


def _is_bullet_text(text: str) -> bool:
    """Check if text is a bullet-point paragraph."""
    return text.startswith("●") or text.startswith("•") or text.startswith("- ")


def _has_multiple_columns_filled(ws, row_idx: int, threshold: int = 2) -> bool:
    """Check if a row has data in multiple columns (tabular, not merged text)."""
    vals = _row_values(ws, row_idx)
    return _non_none_count(vals) >= threshold


def classify_row(
    ws, row_idx: int, catalog: list[MergeInfo], data_width: int
) -> RegionType:
    """Determine the type of a single row."""
    text = _get_row_text(ws, row_idx)

    if not text:
        return RegionType.EMPTY

    is_wide_merge = _is_full_width_merge(row_idx, catalog, data_width)

    # Footnote detection (before other checks — footnotes can be merged too)
    if _is_footnote_row(text):
        return RegionType.FOOTNOTE

    # Full-width merged rows
    if is_wide_merge:
        if _is_bullet_text(text):
            return RegionType.TEXT
        if _is_label_row(text):
            return RegionType.LABEL
        # Short merged text at top of sheet → title
        # Long merged text → treat as text paragraph
        if len(text) > 100:
            return RegionType.TEXT
        return RegionType.TITLE

    # Multi-column row → tabular data
    if _has_multiple_columns_filled(ws, row_idx):
        return RegionType.TABLE

    # Single-column, non-merged text
    if _is_label_row(text):
        return RegionType.LABEL
    if _is_footnote_row(text):
        return RegionType.FOOTNOTE
    if _is_bullet_text(text):
        return RegionType.TEXT

    # Default: if it has content, treat as text
    return RegionType.TEXT


# ---------------------------------------------------------------------------
# Step 3: Group rows into regions
# ---------------------------------------------------------------------------

def detect_regions(ws, catalog: list[MergeInfo]) -> list[Region]:
    """Group consecutive same-type rows into regions."""
    min_col, max_col = ws.min_column, ws.max_column
    data_width = max_col - min_col + 1
    max_row = ws.max_row or 0

    if max_row == 0:
        return []

    regions = []
    current_type = None
    current_start = None

    for row_idx in range(ws.min_row, max_row + 1):
        rtype = classify_row(ws, row_idx, catalog, data_width)

        if rtype == RegionType.EMPTY:
            # Empty rows break regions
            if current_type is not None:
                regions.append(Region(
                    rtype=current_type,
                    start_row=current_start,
                    end_row=row_idx - 1,
                ))
                current_type = None
            continue

        # LABEL always stands alone (even if adjacent to other labels)
        if rtype == RegionType.LABEL:
            if current_type is not None:
                regions.append(Region(
                    rtype=current_type,
                    start_row=current_start,
                    end_row=row_idx - 1,
                ))
            regions.append(Region(rtype=RegionType.LABEL, start_row=row_idx, end_row=row_idx))
            current_type = None
            continue

        # TITLE always stands alone
        if rtype == RegionType.TITLE:
            if current_type is not None:
                regions.append(Region(
                    rtype=current_type,
                    start_row=current_start,
                    end_row=row_idx - 1,
                ))
            regions.append(Region(rtype=RegionType.TITLE, start_row=row_idx, end_row=row_idx))
            current_type = None
            continue

        # Same type continues the region
        if rtype == current_type:
            continue

        # Different type → close previous, start new
        if current_type is not None:
            regions.append(Region(
                rtype=current_type,
                start_row=current_start,
                end_row=row_idx - 1,
            ))

        current_type = rtype
        current_start = row_idx

    # Close last region
    if current_type is not None:
        regions.append(Region(
            rtype=current_type,
            start_row=current_start,
            end_row=max_row,
        ))

    return regions


# ---------------------------------------------------------------------------
# Step 4: Build footnote registries (per-region, since footnotes are inline)
# ---------------------------------------------------------------------------

def build_footnote_registry(ws, region: Region) -> dict[str, str]:
    """Parse a FOOTNOTE region into {marker: explanation} dict."""
    registry = {}
    max_col = ws.max_column

    current_marker = None
    current_text = []

    for row_idx in range(region.start_row, region.end_row + 1):
        line = _get_row_text(ws, row_idx)
        if not line:
            continue

        # Skip keyword-only rows like "Notes:"
        if FOOTNOTE_SECTION_KEYWORDS.match(line) and not FOOTNOTE_EXTRACT_RE.match(line):
            continue

        match = FOOTNOTE_EXTRACT_RE.match(line)
        if match:
            if current_marker:
                registry[current_marker] = " ".join(current_text).strip()
            current_marker = match.group(1).strip()
            current_text = [match.group(2).strip()]
        elif current_marker:
            current_text.append(line)

    if current_marker:
        registry[current_marker] = " ".join(current_text).strip()

    return registry


def collect_footnotes(ws, regions: list[Region]) -> dict[str, str]:
    """
    Collect ALL footnote registries across the sheet into a single dict.
    WARNING: If multiple footnote regions use the same marker (e.g., *),
    use collect_footnotes_per_table() instead.
    """
    combined = {}
    for r in regions:
        if r.rtype == RegionType.FOOTNOTE:
            combined.update(build_footnote_registry(ws, r))
    return combined


def collect_footnotes_per_table(
    ws, regions: list[Region]
) -> dict[int, dict[str, str]]:
    """
    Associate each FOOTNOTE region with the nearest preceding TABLE region.
    Returns {table_region_index: {marker: explanation}}.

    This handles the case where multiple tables use the same marker (e.g., *)
    with different meanings.
    """
    # Map: table_region_idx → list of footnote regions that follow it
    table_footnotes: dict[int, dict[str, str]] = {}
    last_table_idx = None

    for i, r in enumerate(regions):
        if r.rtype == RegionType.TABLE:
            last_table_idx = i
            if i not in table_footnotes:
                table_footnotes[i] = {}
        elif r.rtype == RegionType.FOOTNOTE and last_table_idx is not None:
            fn_reg = build_footnote_registry(ws, r)
            table_footnotes[last_table_idx].update(fn_reg)

    # Also build a "global" registry for TEXT regions (bullet points)
    # that aren't directly after a table. Merge all footnotes for these.
    global_footnotes = {}
    for r in regions:
        if r.rtype == RegionType.FOOTNOTE:
            global_footnotes.update(build_footnote_registry(ws, r))

    table_footnotes["_global"] = global_footnotes
    return table_footnotes


# ---------------------------------------------------------------------------
# Step 5: Resolve footnotes in text
# ---------------------------------------------------------------------------

def resolve_footnotes_in_text(
    text: str,
    registry: dict[str, str],
    inline: bool = True,
) -> tuple[str, dict[str, str]]:
    """Replace footnote markers with their definitions."""
    if not registry or not text:
        return text, {}

    resolved = text
    found = {}

    # Sort markers longest-first to avoid ** matching before ***
    for marker in sorted(registry.keys(), key=len, reverse=True):
        if marker in resolved:
            found[marker] = registry[marker]
            if inline:
                resolved = resolved.replace(marker, f" [Note: {registry[marker]}]")

    return resolved, found


# ---------------------------------------------------------------------------
# Step 6: Detect table headers within a TABLE region
# ---------------------------------------------------------------------------

def _is_number_or_data(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        cleaned = (v.strip().replace(",", "").replace(".", "")
                   .replace("%", "").replace("$", "").replace("€", "")
                   .replace("₺", "").replace(" ", ""))
        if cleaned.isdigit():
            return True
        # Contains digits mixed with text (like "18.398KY kirt") → data
        if any(c.isdigit() for c in v) and len(v) > 5:
            return True
    return False


def detect_table_header(ws, region: Region) -> int | None:
    """
    Find the header row within a TABLE region.

    Strategy: The header row is the first row where cells are short strings.
    We verify it by checking that subsequent rows look structurally different
    (longer text, numeric data, or contain newlines / special patterns).
    """
    min_col, max_col = ws.min_column, ws.max_column

    for row_idx in range(region.start_row, region.end_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value
                for c in range(min_col, max_col + 1)]
        non_none = [v for v in vals if v is not None]

        if len(non_none) < 2:
            continue

        # Check: mostly short, clean strings (no newlines, no bullets)?
        header_score = 0
        for v in non_none:
            s = str(v).strip()
            if isinstance(v, str) and len(s) < 60 and "\n" not in s and not s.startswith("*"):
                header_score += 1

        if header_score / len(non_none) < 0.5:
            continue

        # Verify subsequent rows look like DATA, not more headers.
        # Data rows are: longer text, contain newlines, start with *, have numbers, etc.
        for next_row in range(row_idx + 1, region.end_row + 1):
            next_vals = [ws.cell(row=next_row, column=c).value
                        for c in range(min_col, max_col + 1)]
            next_non_none = [v for v in next_vals if v is not None]

            if not next_non_none:
                continue

            # Any signal that next row is data, not another header?
            is_data = False
            for v in next_non_none:
                s = str(v).strip()
                if _is_number_or_data(v):
                    is_data = True; break
                # Longer text than header cells
                if len(s) > 60:
                    is_data = True; break
                # Contains newlines (multi-value cells)
                if "\n" in str(v):
                    is_data = True; break
                # Starts with * (footnote-style bullet)
                if s.startswith("*"):
                    is_data = True; break

            if is_data:
                return row_idx

            # If next row also looks header-like, try one more row
            continue

        # If we reach here, first row is a candidate but no clear data below
        # Default: treat first row as header if it's the first row of the region
        if row_idx == region.start_row:
            return row_idx

    return None


# ---------------------------------------------------------------------------
# Step 7: Chunk each region type
# ---------------------------------------------------------------------------

def chunk_title_region(ws, region: Region, sheet_name: str) -> list[Chunk]:
    """Title regions become metadata context, not standalone chunks."""
    # Titles are used as context for subsequent regions, not chunked directly
    return []


def chunk_text_region(
    ws, region: Region, context: list[str],
    footnote_registry: dict[str, str], sheet_name: str,
) -> list[Chunk]:
    """Each bullet-point / paragraph row becomes its own chunk."""
    chunks = []
    min_col, max_col = ws.min_column, ws.max_column

    for row_idx in range(region.start_row, region.end_row + 1):
        text = _get_row_text(ws, row_idx)
        if not text:
            continue

        # Resolve footnotes
        text, found_fn = resolve_footnotes_in_text(text, footnote_registry, inline=True)

        # Add hierarchy context
        prefix = " > ".join(context)
        chunk_text = f"[{prefix}] {text}" if prefix else text

        chunks.append(Chunk(
            text=chunk_text,
            metadata={
                "source_type": "excel",
                "sheet_name": sheet_name,
                "row": row_idx,
                "content_type": "text",
                "hierarchy": list(context),
            }
        ))

    return chunks


def chunk_table_region(
    ws, region: Region, context: list[str], label: str | None,
    footnote_registry: dict[str, str], sheet_name: str,
) -> list[Chunk]:
    """Chunk a table region: one chunk per data row with column headers."""
    chunks = []
    min_col, max_col = ws.min_column, ws.max_column

    # Detect header row
    header_row = detect_table_header(ws, region)

    # Build column headers
    col_headers = {}
    if header_row:
        for col in range(min_col, max_col + 1):
            val = ws.cell(row=header_row, column=col).value
            if val:
                # Clean \n from header cells
                col_headers[col] = str(val).replace("\n", " / ").strip()
            else:
                col_headers[col] = f"Column_{col}"
        data_start = header_row + 1
    else:
        for col in range(min_col, max_col + 1):
            col_headers[col] = f"Column_{col}"
        data_start = region.start_row

    # Build full context: hierarchy + label
    full_context = list(context)
    if label:
        full_context.append(label)

    for row_idx in range(data_start, region.end_row + 1):
        parts = []
        for col in range(min_col, max_col + 1):
            val = ws.cell(row=row_idx, column=col).value
            if val is None:
                continue

            col_name = col_headers.get(col, f"Column_{col}")
            cell_text = str(val).replace("\n", " / ").strip()

            # Resolve footnotes in cell text
            if footnote_registry:
                cell_text, _ = resolve_footnotes_in_text(
                    cell_text, footnote_registry, inline=True
                )
            # Also resolve footnotes in column header for this chunk
            if footnote_registry:
                col_name, _ = resolve_footnotes_in_text(
                    col_name, footnote_registry, inline=True
                )

            parts.append(f"{col_name}: {cell_text}")

        if not parts:
            continue

        prefix = " > ".join(full_context)
        row_text = " | ".join(parts)
        chunk_text = f"[{prefix}] {row_text}" if prefix else row_text

        chunks.append(Chunk(
            text=chunk_text,
            metadata={
                "source_type": "excel",
                "sheet_name": sheet_name,
                "row": row_idx,
                "content_type": "table_row",
                "hierarchy": list(full_context),
            }
        ))

    return chunks


# ---------------------------------------------------------------------------
# Step 8: Main pipeline — orchestrate region chunking with context
# ---------------------------------------------------------------------------

def process_sheet(
    ws, catalog: list[MergeInfo], sheet_name: str,
    inline_footnotes: bool = True,
) -> list[Chunk]:
    """Process a single sheet: detect regions, collect footnotes, chunk."""

    # Detect regions
    regions = detect_regions(ws, catalog)

    # Collect all footnotes first (they can be anywhere in the sheet)
    # Use per-table association to handle duplicate markers (e.g., * in multiple tables)
    table_footnotes = collect_footnotes_per_table(ws, regions)
    global_footnotes = table_footnotes.get("_global", {})

    # Build chunks with rolling context
    all_chunks = []
    # context tracks hierarchy as list of (col_span, text) tuples
    # Same col_span = same level → replace, not append
    context_levels: list[tuple[int, str]] = []
    pending_label = None

    def _get_merge_width(row_idx: int) -> int:
        """Get the col_span of the merge at this row, or 0 if not merged."""
        for m in catalog:
            if m.min_row <= row_idx <= m.max_row and m.is_horizontal_header:
                return m.col_span
        return 0

    def _context_strings() -> list[str]:
        return [text for _, text in context_levels]

    for i, region in enumerate(regions):
        if region.rtype == RegionType.TITLE:
            text = _get_row_text(ws, region.start_row)
            if text:
                width = _get_merge_width(region.start_row)
                # Replace existing entry at same width (same hierarchy level)
                # Also remove any narrower entries (sub-sections of old context)
                context_levels = [
                    (w, t) for w, t in context_levels
                    if w > width  # keep only wider (higher-level) entries
                ]
                context_levels.append((width, text.strip()))
                # Sort: widest first (top-level → subsection)
                context_levels.sort(key=lambda x: -x[0])

        elif region.rtype == RegionType.LABEL:
            text = _get_row_text(ws, region.start_row)
            pending_label = text.rstrip(";:").strip() if text else None

        elif region.rtype == RegionType.TEXT:
            chunks = chunk_text_region(
                ws, region, _context_strings(), global_footnotes, sheet_name
            )
            all_chunks.extend(chunks)

        elif region.rtype == RegionType.TABLE:
            # Use table-specific footnotes to avoid marker collisions
            fn_registry = table_footnotes.get(i, {})
            chunks = chunk_table_region(
                ws, region, _context_strings(), pending_label,
                fn_registry, sheet_name,
            )
            all_chunks.extend(chunks)
            pending_label = None  # consumed

        elif region.rtype == RegionType.FOOTNOTE:
            pass  # already collected above

    return all_chunks


def process_excel(
    filepath: str,
    inline_footnotes: bool = True,
    sheets: list[str] | None = None,
) -> list[Chunk]:
    """
    Full auto-detection pipeline for messy Excel files.

    Handles:
    - Multiple tables per sheet
    - Mixed tabular + paragraph content
    - Footnotes between tables (not just at bottom)
    - Bare *, **, *** footnote markers
    - Merged cells (horizontal headers + vertical labels)
    - Subtitle/label rows before tables

    Args:
        filepath: Path to .xlsx file
        inline_footnotes: Inline footnote text into chunks (recommended True)
        sheets: Sheet names to process. None = all sheets.
    """
    wb = load_workbook(filepath, data_only=True)
    sheet_names = sheets or wb.sheetnames
    all_chunks = []

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] Sheet '{sheet_name}' not found, skipping.")
            continue

        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row == 0:
            print(f"[SKIP] Sheet '{sheet_name}' is empty.")
            continue

        # Catalog merges before unmerging
        catalog = catalog_merges(ws)
        unmerge_and_fill(ws, catalog)

        chunks = process_sheet(ws, catalog, sheet_name, inline_footnotes)
        all_chunks.extend(chunks)

        print(
            f"[INFO] Sheet '{sheet_name}': "
            f"{len(chunks)} chunks, {len(catalog)} merges"
        )

    if not all_chunks:
        print("[WARN] No chunks produced from any sheet.")

    return all_chunks


# ---------------------------------------------------------------------------
# Debug utility
# ---------------------------------------------------------------------------

def debug_sheet(filepath: str, sheet_name: str | None = None):
    """Print detected regions for each sheet."""
    wb = load_workbook(filepath, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames

    for sname in sheets:
        ws = wb[sname]
        print(f"\n{'='*70}")
        print(f"Sheet: {sname} | Rows: {ws.min_row}-{ws.max_row}")

        catalog = catalog_merges(ws)
        print(f"Merged ranges: {len(catalog)}")

        unmerge_and_fill(ws, catalog)

        regions = detect_regions(ws, catalog)
        print(f"\nDetected regions ({len(regions)}):")
        for i, r in enumerate(regions):
            text_preview = _get_row_text(ws, r.start_row)[:80]
            print(f"  [{i}] {r.rtype.value:10s} rows {r.start_row}-{r.end_row}: {text_preview}...")

        fn_registry = collect_footnotes(ws, regions)
        if fn_registry:
            print(f"\nFootnotes ({len(fn_registry)}):")
            for marker, text in fn_registry.items():
                print(f"  '{marker}' → {text[:80]}...")

        # Show table header detection
        for i, r in enumerate(regions):
            if r.rtype == RegionType.TABLE:
                header = detect_table_header(ws, r)
                if header:
                    vals = _row_values(ws, header)
                    print(f"\n  Table (rows {r.start_row}-{r.end_row}) header at row {header}: "
                          f"{[v for v in vals if v]}")


if __name__ == "__main__":
    import sys
    import json

    if len(sys.argv) < 2:
        print("Usage: python excel_chunker_v3.py <file.xlsx> [--debug]")
        sys.exit(1)

    filepath = sys.argv[1]

    if "--debug" in sys.argv:
        debug_sheet(filepath)
    else:
        chunks = process_excel(filepath, inline_footnotes=True)
        print(f"\nTotal chunks: {len(chunks)}")
        for i, chunk in enumerate(chunks):
            print(f"\n--- Chunk {i+1} ({chunk.metadata.get('content_type')}) ---")
            print(f"Text: {chunk.text[:200]}")
            print(f"Hierarchy: {chunk.metadata.get('hierarchy')}")
